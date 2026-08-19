from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT="Arial"; NAVY="1F3864"; LIGHT="DDEBF7"; GREEN="E2EFDA"; YELLOW="FFF2CC"
hdr_font=Font(name=FONT,size=10,bold=True,color="FFFFFF")
hdr_fill=PatternFill("solid",fgColor=NAVY)
sub=Font(name=FONT,size=12,bold=True,color=NAVY)
base=Font(name=FONT,size=10)
thin=Side(style="thin",color="BFBFBF"); border=Border(left=thin,right=thin,top=thin,bottom=thin)

wb=load_workbook("Aruba-Prospect-Tracker.xlsx")
R=800

# ---------------- DASHBOARD ----------------
d=wb.create_sheet("Dashboard",1)
d.sheet_view.showGridLines=False
d["B2"]="DASHBOARD"; d["B2"].font=Font(name=FONT,size=18,bold=True,color=NAVY)
d["B3"]="Calculates automatically from the Prospects tab. Review every Friday."
d["B3"].font=Font(name=FONT,size=10,italic=True,color="666666")

def section(row,title):
    d.cell(row,2,title).font=sub

def kv(row,label,formula,fmt=None,note=""):
    d.cell(row,2,label).font=Font(name=FONT,size=10,bold=True)
    c=d.cell(row,4,formula); c.font=base; c.border=border
    c.fill=PatternFill("solid",fgColor=LIGHT)
    c.alignment=Alignment(horizontal="center")
    if fmt: c.number_format=fmt
    if note:
        n=d.cell(row,5,note); n.font=Font(name=FONT,size=9,italic=True,color="666666")

section(5,"THE ISLAND AUDIT")
kv(6,"Businesses logged",f'=COUNTA(Prospects!B4:B{R})','#,##0')
kv(7,"With no website",f'=COUNTIF(Prospects!I4:I{R},"No")','#,##0')
kv(8,"% with no website",f'=IFERROR(D7/D6,0)','0.0%','<-- this is your local statistic. Quote it in every pitch.')
kv(9,"Have a site but not mobile-friendly",f'=COUNTIF(Prospects!J4:J{R},"No")','#,##0','Rebuild prospects')
kv(10,"No Google Business Profile",f'=COUNTIF(Prospects!K4:K{R},"No")','#,##0','Fastest visible win you can sell')

section(12,"PIPELINE")
kv(13,"Not started",f'=COUNTIF(Prospects!U4:U{R},"Not started")','#,##0')
kv(14,"In sequence",f'=COUNTIF(Prospects!U4:U{R},"Sequenced")','#,##0')
kv(15,"Replied",f'=COUNTIF(Prospects!U4:U{R},"Replied")','#,##0')
kv(16,"Meetings booked",f'=COUNTIF(Prospects!U4:U{R},"Meeting booked")','#,##0')
kv(17,"Proposals sent",f'=COUNTIF(Prospects!U4:U{R},"Proposal sent")','#,##0')
kv(18,"WON",f'=COUNTIF(Prospects!U4:U{R},"Won")','#,##0')
kv(19,"Lost",f'=COUNTIF(Prospects!U4:U{R},"Lost")','#,##0')
kv(20,"Unsubscribed",f'=COUNTIF(Prospects!U4:U{R},"Unsubscribed")','#,##0','Never contact again')
d.cell(18,2).font=Font(name=FONT,size=10,bold=True,color="006100")
d.cell(18,4).fill=PatternFill("solid",fgColor=GREEN)

section(22,"REVENUE")
kv(23,"Build fees won",f'=SUMIF(Prospects!U4:U{R},"Won",Prospects!V4:V{R})','$#,##0')
kv(24,"Monthly recurring won",f'=SUMIF(Prospects!U4:U{R},"Won",Prospects!W4:W{R})','$#,##0')
kv(25,"Annual recurring run-rate",'=D24*12','$#,##0')
kv(26,"Total booked (yr 1)",'=D23+D25','$#,##0')
kv(27,"Avg build fee per client",'=IFERROR(D23/D18,0)','$#,##0','If under $500 you are discounting too hard')

section(29,"CONVERSION")
kv(30,"Contacted (all touched)",f'=D6-D13','#,##0')
kv(31,"Reply rate",'=IFERROR((D15+D16+D17+D18+D19)/D30,0)','0.0%','Benchmark 3.4%. Good is 6%+.')
kv(32,"Contacted -> meeting",'=IFERROR((D16+D17+D18)/D30,0)','0.0%','Target 2-3%')
kv(33,"Meeting -> won",'=IFERROR(D18/(D16+D17+D18+D19),0)','0.0%','Target 30%. With a demo built, 50%+.')

section(35,"BY PRIORITY TIER")
for i,(lbl,t) in enumerate([("Tier A","Tier A"),("Tier B","Tier B"),("Tier C","Tier C")]):
    kv(36+i,lbl+" logged",f'=COUNTIF(Prospects!L4:L{R},"{t}")','#,##0')
kv(39,"Tier A won",f'=COUNTIFS(Prospects!L4:L{R},"Tier A",Prospects!U4:U{R},"Won")','#,##0')

section(41,"BY SECTOR — where you are actually winning")
d.cell(42,2,"Sector").font=hdr_font; d.cell(42,2).fill=hdr_fill
d.cell(42,3,"Logged").font=hdr_font; d.cell(42,3).fill=hdr_fill
d.cell(42,4,"No site").font=hdr_font; d.cell(42,4).fill=hdr_fill
d.cell(42,5,"Won").font=hdr_font; d.cell(42,5).fill=hdr_fill
for cc in (2,3,4,5):
    d.cell(42,cc).alignment=Alignment(horizontal="center")
SECTORS=["Tours & Watersports","Guesthouse / Rental","Restaurant / Bar","Food Truck",
         "Trades / Contractor","Auto / Car Rental","Professional Services","Health / Clinic",
         "Retail / Boutique","Salon / Barber","Gym / Fitness","Events / Wedding","Other"]
for i,s in enumerate(SECTORS):
    rr=43+i
    d.cell(rr,2,s).font=base; d.cell(rr,2).border=border
    d.cell(rr,3,f'=COUNTIF(Prospects!$C$4:$C${R},$B{rr})').font=base
    d.cell(rr,4,f'=COUNTIFS(Prospects!$C$4:$C${R},$B{rr},Prospects!$I$4:$I${R},"No")').font=base
    d.cell(rr,5,f'=COUNTIFS(Prospects!$C$4:$C${R},$B{rr},Prospects!$U$4:$U${R},"Won")').font=base
    for cc in (3,4,5):
        d.cell(rr,cc).border=border; d.cell(rr,cc).alignment=Alignment(horizontal="center")
        d.cell(rr,cc).number_format='#,##0'
tot=43+len(SECTORS)
d.cell(tot,2,"TOTAL").font=Font(name=FONT,size=10,bold=True)
for cc,col in ((3,"C"),(4,"D"),(5,"E")):
    c=d.cell(tot,cc,f'=SUM({col}43:{col}{tot-1})')
    c.font=Font(name=FONT,size=10,bold=True); c.border=border
    c.alignment=Alignment(horizontal="center"); c.number_format='#,##0'

d.column_dimensions["A"].width=2
d.column_dimensions["B"].width=34
d.column_dimensions["C"].width=11
d.column_dimensions["D"].width=13
d.column_dimensions["E"].width=52

# ---------------- SECTORS ----------------
s=wb.create_sheet("Sectors")
s.sheet_view.showGridLines=False
s["B2"]="TARGET SECTORS — ranked, with Google Maps search terms"
s["B2"].font=Font(name=FONT,size=14,bold=True,color=NAVY)
s["B3"]="Work top to bottom. Tier A first — best ROI story, easiest close."
s["B3"].font=Font(name=FONT,size=10,italic=True,color="666666")
heads=["Tier","Sector","Why they buy","Search terms for Google Maps","Est. % with no site"]
widths=[8,24,52,46,16]
for i,(h,w) in enumerate(zip(heads,widths),start=2):
    c=s.cell(5,i,h); c.font=hdr_font; c.fill=hdr_fill; c.border=border
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    s.column_dimensions[get_column_letter(i)].width=w
s.row_dimensions[5].height=28
data=[
 ("A","Tours & Watersports","Pays Viator/TripAdvisor 20-30% commission. One direct booking pays the year.","snorkel tour aruba / UTV tour / sunset sail / dive shop / kitesurf / fishing charter","High"),
 ("A","Guesthouse / Rental","Pays Booking.com & Airbnb 15-18%. Direct booking form is pure margin.","guesthouse aruba / apartment rental aruba / vacation rental / B&B","High"),
 ("A","Restaurant / Bar","Menu is a Facebook photo — Google can't read it. Tourists search 'restaurants near me'.","restaurant oranjestad / bar noord / seafood aruba / local food aruba","35-42%"),
 ("A","Food Truck","Location + hours change constantly. A live page beats a buried FB post.","food truck aruba / snack aruba / truck di cuminda","High"),
 ("B","Trades / Contractor","Highest no-website rate anywhere. Locals search at 11pm when it breaks. Best retention.","ac repair aruba / plumber aruba / electrician / pool service / roofing / solar","45-56%"),
 ("B","Auto / Car Rental","Independents lose every tourist search to Hertz and Avis.","car rental aruba / scooter rental / mechanic aruba / tyre shop / body shop","38-51%"),
 ("B","Professional Services","Credibility is the product. 75% judge a business by its site. Highest ability to pay.","accountant aruba / lawyer / notary / insurance broker / translator","15-25%"),
 ("B","Health / Clinic","Same credibility argument. Slower sale, better client.","dentist aruba / physio / vet aruba / clinic","20-30%"),
 ("C","Retail / Boutique","Lower ticket, but visual and good for your portfolio.","boutique aruba / gift shop / mini market / jewellery","40-48%"),
 ("C","Salon / Barber","Price-sensitive and high churn — take them, don't chase them.","salon aruba / barber aruba / nails / spa","38-44%"),
 ("C","Gym / Fitness","Class schedules and pricing pages sell themselves.","gym aruba / crossfit / yoga aruba","Medium"),
 ("C","Events / Wedding","Tourist-driven and high-ticket, but seasonal.","wedding planner aruba / photographer aruba / event rental","Medium"),
]
for i,row in enumerate(data):
    rr=6+i
    fill=GREEN if row[0]=="A" else (YELLOW if row[0]=="B" else None)
    for j,v in enumerate(row,start=2):
        c=s.cell(rr,j,v); c.font=base; c.border=border
        c.alignment=Alignment(wrap_text=True,vertical="top")
        if fill and j==2: c.fill=PatternFill("solid",fgColor=fill)
        if j==2: c.alignment=Alignment(horizontal="center",vertical="center")
    s.row_dimensions[rr].height=34
note=s.cell(6+len(data)+2,2,"Skip: large resorts, AHATA hotel members, franchise/chain outlets, anything where the person you're emailing isn't the owner.")
note.font=Font(name=FONT,size=10,italic=True,color="666666")
s.cell(6+len(data)+3,2,"'Est. % with no site' figures are global industry benchmarks, not Aruba-measured. Replace them with your own Island Audit numbers as you build the list.").font=Font(name=FONT,size=9,italic=True,color="666666")
s.column_dimensions["A"].width=2

# ---------------- SEND LOG ----------------
L=wb.create_sheet("Send Log")
L.sheet_view.showGridLines=False
L["B2"]="DAILY SEND LOG"; L["B2"].font=Font(name=FONT,size=14,bold=True,color=NAVY)
L["B3"]="Log every sending day. Blue = you type it. Grey = calculated."
L["B3"].font=Font(name=FONT,size=10,italic=True,color="666666")
sh=["Date","Sent","Bounced","Replies","Positive","Meetings booked","Bounce %","Reply %"]
sw=[13,10,11,10,10,17,11,11]
for i,(h,w) in enumerate(zip(sh,sw),start=2):
    c=L.cell(5,i,h); c.font=hdr_font; c.fill=hdr_fill; c.border=border
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    L.column_dimensions[get_column_letter(i)].width=w
L.row_dimensions[5].height=28
LR=205
ex=["2026-09-01",25,0,1,1,0]
for j,v in enumerate(ex,start=2):
    c=L.cell(6,j,v); c.font=Font(name=FONT,size=10,color="0000FF"); c.border=border
    c.alignment=Alignment(horizontal="center")
for rr in range(6,LR+1):
    for j in range(2,8):
        c=L.cell(rr,j)
        if rr>6: c.font=Font(name=FONT,size=10,color="0000FF")
        c.border=border; c.alignment=Alignment(horizontal="center")
    bc=L.cell(rr,8,f'=IFERROR(D{rr}/C{rr},"")'); bc.number_format='0.0%'
    rc=L.cell(rr,9,f'=IFERROR(E{rr}/C{rr},"")'); rc.number_format='0.0%'
    for c in (bc,rc):
        c.font=base; c.border=border
        c.alignment=Alignment(horizontal="center")
        c.fill=PatternFill("solid",fgColor="F2F2F2")
tr=LR+2
L.cell(tr,2,"TOTAL").font=Font(name=FONT,size=10,bold=True)
for j,col in ((3,"C"),(4,"D"),(5,"E"),(6,"F"),(7,"G")):
    c=L.cell(tr,j,f'=SUM({col}6:{col}{LR})'); c.font=Font(name=FONT,size=10,bold=True)
    c.border=border; c.alignment=Alignment(horizontal="center"); c.number_format='#,##0'
c=L.cell(tr,8,f'=IFERROR(D{tr}/C{tr},"")'); c.number_format='0.0%'
c.font=Font(name=FONT,size=10,bold=True); c.border=border; c.alignment=Alignment(horizontal="center")
c=L.cell(tr,9,f'=IFERROR(E{tr}/C{tr},"")'); c.number_format='0.0%'
c.font=Font(name=FONT,size=10,bold=True); c.border=border; c.alignment=Alignment(horizontal="center")
L.cell(tr+2,2,"STOP SENDING if bounce % goes over 3% or you get spam complaints. Verify the list, then restart slowly.").font=Font(name=FONT,size=10,bold=True,color="C00000")
L.cell(tr+3,2,"Safe volume: 20-30/day for Aruba. Never exceed 100/day per mailbox; over 150/day raises spam rates 43%.").font=Font(name=FONT,size=9,italic=True,color="666666")
L.column_dimensions["A"].width=2

wb.save("Aruba-Prospect-Tracker.xlsx")
print("done")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FONT = "Arial"
NAVY = "1F3864"; BLUE = "2E5E8C"; LIGHT = "DDEBF7"; YELLOW = "FFFF00"
GREY = "F2F2F2"; GREEN = "E2EFDA"

hdr_font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor=NAVY)
sub_font = Font(name=FONT, size=11, bold=True, color=NAVY)
base_font = Font(name=FONT, size=10)
input_font = Font(name=FONT, size=10, color="0000FF")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ---------------- START HERE ----------------
ws = wb.active
ws.title = "Start Here"
ws.sheet_view.showGridLines = False
ws["B2"] = "ARUBA PROSPECT TRACKER"
ws["B2"].font = Font(name=FONT, size=18, bold=True, color=NAVY)
ws["B3"] = "Website sales pipeline — built August 2026"
ws["B3"].font = Font(name=FONT, size=10, italic=True, color="666666")

rows = [
    ("", ""),
    ("HOW TO USE THIS FILE", ""),
    ("1.", "Fill in the 'Prospects' tab. Blue text = cells you type in. Row 4 is an example — delete it once you start."),
    ("2.", "Log every business you find, WITH or WITHOUT a website. The ones with no site are your prospects; the full count is your Island Audit statistic."),
    ("3.", "Use the dropdowns in Sector, District, Priority, Channel and Status — the Dashboard only counts correctly if you do."),
    ("4.", "Update Status as you go. Log each touch date in the Touch 1-5 columns."),
    ("5.", "The 'Dashboard' tab calculates itself. Check it every Friday."),
    ("6.", "'Sectors' tab = your target priority list and the search terms to use on Google Maps."),
    ("7.", "'Send Log' tab = daily email volume and deliverability. Watch the bounce rate — over 3% means stop and fix."),
    ("", ""),
    ("THE TWO NUMBERS THAT MATTER MOST", ""),
    ("Reply rate", "Under 2% after 300 emails = your first line isn't specific enough, or you have no demo link, or you're in spam. Check in that order."),
    ("Bounce rate", "Over 3% = stop sending immediately. Verify your list before you damage the sending domain."),
    ("", ""),
    ("PRIORITY TIERS", ""),
    ("Tier A", "Tours, watersports, guesthouses, vacation rentals, non-strip restaurants. Build a real demo before emailing."),
    ("Tier B", "Trades, contractors, auto services, professional services. Generic sector demo with their name dropped in."),
    ("Tier C", "Retail, salons, gyms, everyone else. No demo — use a Google search screenshot instead."),
    ("", ""),
    ("STATUS MEANINGS", ""),
    ("Not started", "Logged, not yet contacted"),
    ("Sequenced", "In the 5-touch email sequence"),
    ("Replied", "They responded — reply within 2 hours"),
    ("Meeting booked", "Get off email and into a room. This is the goal of every email."),
    ("Proposal sent", "Showed them the three tiers"),
    ("Won", "Signed. Log deal value + monthly."),
    ("Lost", "No, for now. Revisit in 6 months."),
    ("Unsubscribed", "Asked to be removed. Never contact again."),
]
r = 5
for a, b in rows:
    ws.cell(r, 2, a); ws.cell(r, 3, b)
    if a and not b:
        ws.cell(r, 2).font = sub_font
    else:
        ws.cell(r, 2).font = Font(name=FONT, size=10, bold=True)
        ws.cell(r, 3).font = base_font
        ws.cell(r, 3).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 17
ws.column_dimensions["C"].width = 105

# ---------------- PROSPECTS ----------------
p = wb.create_sheet("Prospects")
cols = [
    ("ID", 6), ("Business Name", 26), ("Sector", 20), ("District", 14),
    ("Owner Name", 16), ("Phone / WhatsApp", 17), ("Email", 26),
    ("Facebook / IG", 24), ("Has Website?", 12), ("Mobile OK?", 11),
    ("Google Profile?", 14), ("Priority", 10), ("Demo Built?", 12),
    ("Demo Link", 24), ("Channel", 14),
    ("Touch 1", 11), ("Touch 2", 11), ("Touch 3", 11), ("Touch 4", 11), ("Touch 5", 11),
    ("Status", 16), ("Deal $", 10), ("Monthly $", 11), ("Notes", 40),
]
p["A1"] = "PROSPECTS — blue text columns are for you to fill in"
p["A1"].font = Font(name=FONT, size=12, bold=True, color=NAVY)
p["A2"] = "Row 4 is an example. Delete it when you start."
p["A2"].font = Font(name=FONT, size=9, italic=True, color="666666")

for i, (name, w) in enumerate(cols, start=1):
    c = p.cell(3, i, name)
    c.font = hdr_font; c.fill = hdr_fill; c.border = border
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    p.column_dimensions[get_column_letter(i)].width = w
p.row_dimensions[3].height = 30
p.freeze_panes = "C4"

example = [1, "Playa Snorkel Tours", "Tours & Watersports", "Noord", "Ricardo",
           "+297 593 1234", "info@playasnorkel.aw", "fb.com/playasnorkel", "No", "N/A",
           "No", "Tier A", "Yes", "demo.mysite.com/playa", "Email",
           "2026-09-01", "2026-09-05", "2026-09-09", "", "", "Replied", 500, 20,
           "Pays Viator 20% — lead with the commission argument"]
for i, v in enumerate(example, start=1):
    c = p.cell(4, i, v)
    c.font = input_font; c.border = border
    c.alignment = Alignment(vertical="top", wrap_text=(i == 24))

MAXROW = 800
for rr in range(5, MAXROW + 1):
    for i in range(1, len(cols) + 1):
        c = p.cell(rr, i)
        c.font = input_font; c.border = border
p["V4"].number_format = '$#,##0'; p["W4"].number_format = '$#,##0'
for rr in range(4, MAXROW + 1):
    p.cell(rr, 22).number_format = '$#,##0'
    p.cell(rr, 23).number_format = '$#,##0'

SECTORS = ["Tours & Watersports","Guesthouse / Rental","Restaurant / Bar","Food Truck",
           "Trades / Contractor","Auto / Car Rental","Professional Services","Health / Clinic",
           "Retail / Boutique","Salon / Barber","Gym / Fitness","Events / Wedding","Other"]
DISTRICTS = ["Oranjestad","Noord","Palm Beach","Eagle Beach","Malmok","San Nicolas",
             "Santa Cruz","Savaneta","Paradera","Piedra Plat","Other"]
YN = ["Yes","No","Unknown","N/A"]
TIERS = ["Tier A","Tier B","Tier C"]
CHANNELS = ["Email","Messenger","Instagram","WhatsApp","Walk-in"]
STATUS = ["Not started","Sequenced","Replied","Meeting booked","Proposal sent","Won","Lost","Unsubscribed"]

def add_dv(col, options):
    dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True)
    p.add_data_validation(dv)
    dv.add(f"{col}4:{col}{MAXROW}")

add_dv("C", SECTORS); add_dv("D", DISTRICTS)
add_dv("I", YN); add_dv("J", YN); add_dv("K", YN); add_dv("M", YN)
add_dv("L", TIERS); add_dv("O", CHANNELS); add_dv("U", STATUS)

wb.save("Aruba-Prospect-Tracker.xlsx")
print("base saved")
